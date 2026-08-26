# -*- coding: utf-8 -*-
"""Plan de Medios Dale Coopeuch — replica la estructura del plan de Diplas."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/meta-mcp/propuesta-dale-coopeuch/02_Plan_de_Medios_Dale_Coopeuch_Mes_Inicial.xlsx"

BLACK   = "FF000000"
GRAY_H  = "FF999999"
GRAY_D  = "FF3A3838"
WHITE   = "FFFFFFFF"
ORANGE  = "FFE57000"
LIGHT   = "FFF2F2F2"

thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ─────────────────────────────── HOJA 1: PLAN DE MEDIOS ───────────────────────
ws = wb.active
ws.title = "Mes Inicial - Propuesta"

# Cabecera (mismo layout que Diplas: D/E filas 1-4)
ws.merge_cells("B1:C4")
hdr = [("AGENCIA", "Intothecom"), ("CLIENTE", "Dale Coopeuch"),
       ("ASUNTO", "Propuesta Plan de Medios"), ("PERIODO", "Mes Inicial")]
for i, (k, v) in enumerate(hdr, start=1):
    c = ws.cell(row=i, column=4, value=k)
    c.font = Font(name="Calibri", size=12, bold=True)
    c.fill = PatternFill("solid", fgColor=WHITE)
    c2 = ws.cell(row=i, column=5, value=v)
    c2.font = Font(name="Calibri", size=12)
    c2.fill = PatternFill("solid", fgColor=WHITE)

# Título de bloque
ws.merge_cells("A7:L7")
t = ws["A7"]; t.value = "Plan de medios"
t.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=BLACK)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[7].height = 22

COLS = ["Canal", "Tipo de campaña", "Campaña", "Anuncio", "Fecha Inicio", "Fecha Fin",
        "Segmento", "Formato", "Medidas del Contenido", "Ubicación", "Objetivo",
        "Inversión mensual"]
for j, h in enumerate(COLS, start=1):
    c = ws.cell(row=8, column=j, value=h)
    c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=GRAY_D if j == 12 else GRAY_H)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[8].height = 32

AO = "Always on\nSin término"
PD = "Por Definir"

ROWS = [
 ["Google Ads", "Search | Marca",
  "Cobertura de búsquedas de marca para proteger el término propio y capturar la demanda ya "
  "generada, con objetivo de conversión principal \"conversación calificada\".",
  "Anuncios de búsqueda sobre términos de marca que dirigen a la landing de Crédito Digital "
  "con CTA directo a WhatsApp.",
  PD, AO, "Búsquedas de marca y variantes. Sin señales de audiencia adicionales.",
  "Texto", "Títulos 30 car.\nDescripciones 90 car.", "SERP",
  "Conversación calificada", "$200.000 - $300.000"],

 ["Google Ads", "Search | Genérica Crédito de Consumo",
  "Captura de demanda de alta intención sobre crédito de consumo digital, priorizando términos "
  "con intención de contratación por sobre términos informacionales.",
  "Anuncios de búsqueda enfocados en el diferencial de proceso 100% digital sin trámites "
  "presenciales, derivando a conversación asistida.",
  PD, AO, "Señales de audiencia por intención de compra en servicios financieros.",
  "Texto", "Títulos 30 car.\nDescripciones 90 car.", "SERP",
  "Conversación calificada", "$650.000 - $1.100.000"],

 ["Google Ads", "Search | Nicho Inclusión Financiera",
  "Cobertura del nicho desatendido de trabajadores independientes y personas sin liquidación "
  "de sueldo, apalancando la evaluación con múltiples fuentes de ingreso como diferenciador.",
  "Anuncios de búsqueda dirigidos a independientes, honorarios y trabajadores informales, "
  "comunicando la evaluación inteligente e inclusiva.",
  PD, AO, "Términos de nicho. Exclusión de términos de crédito hipotecario y automotriz.",
  "Texto", "Títulos 30 car.\nDescripciones 90 car.", "SERP",
  "Conversación calificada", "$350.000 - $600.000"],

 ["Meta Ads", "CTWA | Prospección Broad",
  "Prospección de mercado abierto con campañas Click to WhatsApp, donde el agente IA levanta "
  "información, califica y deriva al prospecto. Segmentación amplia según recomendación vigente "
  "de la plataforma.",
  "Piezas gráficas y video que comunican el crédito 100% digital, con CTA de inicio de "
  "conversación por WhatsApp.",
  PD, AO, "Broad con Advantage+ Audience. Restricción obligatoria 18+.",
  "Imagen - Video", "Imagen 1:1\nImagen 4:5\nVideo 9:16\nVideo 1:1",
  "Feed | Stories | Reels", "Conversación calificada", "$850.000 - $1.400.000"],

 ["Meta Ads", "CTWA | Ángulo Inclusión",
  "Campaña dedicada al ángulo de inclusión financiera para independientes, testeando el "
  "diferenciador de evaluación con múltiples fuentes de ingreso contra el mensaje genérico.",
  "Piezas centradas en el perfil independiente, con testimonios y explicación del proceso "
  "de evaluación.",
  PD, AO, "Broad con señales de intereses de emprendimiento y trabajo independiente.",
  "Imagen - Video", "Imagen 1:1\nImagen 4:5\nVideo 9:16\nVideo 1:1",
  "Feed | Stories | Reels", "Conversación calificada", "$600.000 - $1.000.000"],

 ["Meta Ads", "Retargeting | Recuperación",
  "Recuperación de usuarios que interactuaron con los anuncios o iniciaron conversación sin "
  "completar la calificación, complementando los flujos automatizados de WhatsApp y email.",
  "Piezas de recordatorio y resolución de objeciones frecuentes del proceso de solicitud.",
  PD, AO, "Audiencias personalizadas de interacción con anuncios y con la cuenta de WhatsApp.",
  "Imagen - Video", "Imagen 1:1\nImagen 4:5\nVideo 9:16",
  "Feed | Stories | Reels", "Conversación calificada", "$350.000 - $600.000"],
]

r = 9
for row in ROWS:
    for j, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        if j in (1, 2):
            c.font = Font(name="Arial", size=10, bold=True)
    ws.row_dimensions[r].height = 92
    r += 1

# Totales
tr = r
ws.cell(row=tr, column=11, value="Total Inversión").font = Font(name="Arial", size=11, bold=True)
ws.cell(row=tr, column=11).alignment = Alignment(horizontal="center")
tc = ws.cell(row=tr, column=12, value="$3.000.000 - $5.000.000")
tc.font = Font(name="Arial", size=11, bold=True)
tc.alignment = Alignment(horizontal="center")
tc.fill = PatternFill("solid", fgColor=LIGHT)
tc.border = BORDER

# Consideraciones
cr = tr + 2
ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=8)
cc = ws.cell(row=cr, column=1, value="Consideraciones")
cc.font = Font(name="Arial", size=11, bold=True)
cc.fill = PatternFill("solid", fgColor=WHITE)

NOTAS = [
 "*La distribución de inversión por plataforma puede variar según el rendimiento observado durante el transcurso de las campañas.",
 "*Los montos indicados son una recomendación de la agencia. La inversión puede ajustarse según el presupuesto definido por la marca.",
 "*El presupuesto de inversión en campañas es aparte del valor del servicio de agencia y se paga directamente a cada plataforma.",
 "*Estimación de resultados sujeta a las métricas reales que se observen al implementar. Ver hoja \"Modelo de Inversión\".",
 "*Requisito previo al lanzamiento en Google Ads: publicar en la landing pública los plazos mínimo y máximo de pago, la CAE máxima y un ejemplo representativo del costo total del crédito.",
 "*Las campañas de crédito exigen segmentación 18+ y pueden requerir verificación del anunciante con acreditación regulatoria.",
]
for i, n in enumerate(NOTAS):
    rr = cr + 1 + i
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
    c = ws.cell(row=rr, column=1, value=n)
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")

widths = {"A": 16, "B": 26, "C": 42, "D": 40, "E": 13, "F": 14, "G": 30,
          "H": 15, "I": 20, "J": 22, "K": 22, "L": 22}
for k, v in widths.items():
    ws.column_dimensions[k].width = v
ws.sheet_view.showGridLines = False
print("Hoja 1 lista")

# ─────────────────────────── HOJA 2: MODELO DE INVERSIÓN ──────────────────────
ws2 = wb.create_sheet("Modelo de Inversión")
ws2.sheet_view.showGridLines = False

def block(ws, row, text, span=6, fill=BLACK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

ws2["A1"] = "Modelo de estimación de inversión — Dale Coopeuch"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True)
ws2.merge_cells("A2:C2")
ws2["A2"] = ("Modelo editable. Las celdas naranjas son supuestos: al modificarlas se recalcula "
             "todo. Se reemplazan por métricas reales una vez implementadas las campañas.")
ws2["A2"].font = Font(name="Calibri", size=10, italic=True, color="FF595959")
ws2["A2"].alignment = Alignment(wrap_text=False)

block(ws2, 4, "SUPUESTOS EDITABLES")
sup = [
 ("CPC Google Search (CLP)", 2800, "Banda triangulada $2.400 - $3.200. Ver hoja Benchmarks."),
 ("CPC Meta CTWA (CLP)", 850, "Derivado del benchmark de tráfico convertido. Supuesto más débil: validar."),
 ("Clic → conversación iniciada (Google)", 0.12, "Landing con CTA directo a WhatsApp."),
 ("Clic → conversación iniciada (Meta CTWA)", 0.25, "El clic abre WhatsApp directamente."),
 ("Conversación → lead calificado (IA)", 0.35, "Calificación por el agente IA antes de derivar."),
 ("Lead calificado → curse", 0.20, "DEPENDE DE LA POLÍTICA DE RIESGO DEL CLIENTE. A definir."),
 ("Inversión mensual total (CLP)", 3500000, "Escenario medio de la banda propuesta."),
 ("Participación Google", 0.40, "El 60% restante se asigna a Meta."),
]
r = 5
for k, v, nota in sup:
    ws2.cell(row=r, column=1, value=k).font = Font(name="Calibri", size=11, bold=True)
    c = ws2.cell(row=r, column=2, value=v)
    c.font = Font(name="Calibri", size=11, bold=True)
    c.fill = PatternFill("solid", fgColor="FFFDE9D9")
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")
    c.number_format = '#,##0' if v > 1 else '0.0%'
    ws2.cell(row=r, column=3, value=nota).font = Font(name="Calibri", size=10, color="FF595959")
    r += 1

block(ws2, 14, "RESULTADO PROYECTADO")
res = [
 ("Inversión Google (CLP)", "=B11*B12", '#,##0'),
 ("Inversión Meta (CLP)", "=B11*(1-B12)", '#,##0'),
 ("Clics Google", "=B15/B5", '#,##0'),
 ("Clics Meta", "=B16/B6", '#,##0'),
 ("Conversaciones iniciadas", "=B17*B7+B18*B8", '#,##0'),
 ("Leads calificados", "=B19*B9", '#,##0'),
 ("Costo por lead calificado (CLP)", "=B11/B20", '#,##0'),
 ("Cursos estimados", "=B20*B10", '#,##0'),
 ("Costo por curse (CLP)", "=B11/B22", '#,##0'),
]
r = 15
for k, f, fmt in res:
    ws2.cell(row=r, column=1, value=k).font = Font(name="Calibri", size=11, bold=True)
    c = ws2.cell(row=r, column=2, value=f)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=GRAY_D)
    c.number_format = fmt
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
    r += 1

ws2.cell(row=25, column=1, value=(
 "Nota metodológica: el modelo se construye de abajo hacia arriba a partir de benchmarks públicos "
 "de la industria financiera, no de un promedio de inversión de mercado — esa cifra no existe de "
 "forma pública y confiable a nivel de anunciante individual. La tasa de aprobación a curse es la "
 "única variable que no puede estimarse desde fuentes externas: depende de la política de riesgo "
 "de Coopeuch y debe ser aportada por el cliente."
)).font = Font(name="Calibri", size=10, italic=True, color="FF595959")
ws2.merge_cells("A25:C28")
ws2["A25"].alignment = Alignment(wrap_text=True, vertical="top")
for _rh in range(25, 29):
    ws2.row_dimensions[_rh].height = 26

for col, w in {"A": 40, "B": 18, "C": 62, "D": 12, "E": 12, "F": 12}.items():
    ws2.column_dimensions[col].width = w

# ─────────────────────────── HOJA 3: BENCHMARKS ───────────────────────────────
ws3 = wb.create_sheet("Benchmarks de Referencia")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "Benchmarks de referencia — industria financiera"
ws3["A1"].font = Font(name="Calibri", size=14, bold=True)

hd = ["Métrica", "Valor", "Alcance", "Fuente", "Fecha"]
for j, h in enumerate(hd, start=1):
    c = ws3.cell(row=3, column=j, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=GRAY_D)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

BM = [
 ("CPC Search finanzas y seguros", "USD 3,46", "EE.UU.", "WordStream Google Ads Benchmarks", "2026"),
 ("CPC Search finanzas y banca", "USD 3,08", "Global", "Digital Applied / WordStream Q1", "2026"),
 ("CTR Search finanzas y seguros", "8,3%", "EE.UU.", "WordStream Google Ads Benchmarks", "2026"),
 ("Tasa de conversión Search finanzas", "2,5% - 4,72%", "EE.UU. / Global", "WordStream / Digital Applied", "2026"),
 ("CPC servicios financieros", "$1.500 - $5.000 CLP", "Chile", "Reportes de agencias locales", "2026"),
 ("CPC financiamiento y seguros", "$2.000 - $3.000 CLP", "Chile", "Reportes de agencias locales", "2026"),
 ("CPC Meta finanzas (conversión)", "USD 3,77", "EE.UU.", "Benchmarks Meta Ads por industria", "2026"),
 ("CPC Meta finanzas (tráfico)", "USD 1,22", "EE.UU.", "Benchmarks Meta Ads por industria", "2026"),
 ("Click rate email finanzas y banca", "3,4%", "Global", "Benchmarks email por industria", "2026"),
 ("Click rate flujos vs campañas", "5,58% vs 1,69%", "Global", "Klaviyo (183.000 cuentas)", "2026"),
 ("Ingreso de flujos sobre total email", "41% con 5,3% de envíos", "Global", "Klaviyo (183.000 cuentas)", "2026"),
 ("Engagement Instagram serv. financieros", "0,26% - 0,67%", "Global", "Rival IQ", "2026"),
 ("Engagement TikTok serv. financieros", "1,9%", "Global", "Rival IQ", "2026"),
 ("Cartera consumo sobre total cooperativas", "68,74%", "Chile", "CMF", "ene-2026"),
 ("Crecimiento real cartera consumo cooperativas", "4,77% a 12 meses", "Chile", "CMF", "ene-2026"),
 ("Share of Investment digital", "50,1%", "Chile", "AAM / Admetricks / IAB Chile", "may-2026"),
 ("Participación social sobre inversión digital", "37,0%", "Chile", "AAM / Admetricks / IAB Chile", "may-2026"),
 ("Participación search sobre inversión digital", "31,3%", "Chile", "AAM / Admetricks / IAB Chile", "may-2026"),
 ("Valor UF de referencia", "$40.867,18 CLP", "Chile", "Banco Central", "26-ago-2026"),
 ("Valor dólar observado de referencia", "$911,43 CLP", "Chile", "Banco Central", "26-ago-2026"),
]
r = 4
for row in BM:
    for j, v in enumerate(row, start=1):
        c = ws3.cell(row=r, column=j, value=v)
        c.font = Font(name="Calibri", size=10)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=(j == 1))
    r += 1

ws3.cell(row=r+1, column=1, value=(
 "Los benchmarks internacionales se usan como referencia de la relación entre métricas, no como "
 "valor absoluto trasladable a Chile. Los valores en pesos surgen de triangular la fuente local "
 "contra la internacional convertida al tipo de cambio de referencia."
)).font = Font(name="Calibri", size=10, italic=True, color="FF595959")
ws3.merge_cells(start_row=r+1, start_column=1, end_row=r+2, end_column=5)
ws3.cell(row=r+1, column=1).alignment = Alignment(wrap_text=True, vertical="top")

for col, w in {"A": 42, "B": 22, "C": 18, "D": 40, "E": 14}.items():
    ws3.column_dimensions[col].width = w

# ─────────────────── Configuración de impresión / exportación a PDF ───────────
for sheet, cols in ((ws, "L"), (ws2, "C"), (ws3, "E")):
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = sheet.page_margins.right = 0.3
    sheet.page_margins.top = sheet.page_margins.bottom = 0.4
    sheet.print_options.horizontalCentered = True
ws.print_area = f"A1:L{cr + len(NOTAS)}"
ws2.print_area = "A1:C28"
ws3.print_area = f"A1:E{r + 2}"
ws.freeze_panes = "A9"

wb.save(OUT)
print("OK →", OUT)
